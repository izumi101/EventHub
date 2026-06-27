"""Attendee export for organizers — a styled .xlsx workbook.

Real Excel format (not CSV) so it opens with proper columns regardless of the
user's locale/delimiter settings, with a bold frozen header, auto-sized
columns, and an auto-filter.
"""
from io import BytesIO

from django.http import HttpResponse

from common.choices import RegistrationStatusChoices

HEADERS = [
    'Name', 'Username', 'Email', 'Ticket type', 'Status', 'Checked in',
    'Seat', 'Zone', 'Promo code', 'Amount paid', 'Registered at', 'Ticket ID',
]


def _rows(event):
    regs = (
        event.registrations
        .exclude(status=RegistrationStatusChoices.CANCELLED)
        .select_related('user', 'seat', 'promo_code', 'payment', 'ticket_type')
        .order_by('registered_at')
    )
    for r in regs:
        u = r.user
        full_name = f'{u.first_name} {u.last_name}'.strip() or u.username
        tier = r.ticket_type.name if r.ticket_type_id else ''
        seat = f'Row {r.seat.row} Seat {r.seat.col}' if r.seat else ''
        zone = r.seat.price_zone if r.seat else ''
        promo = r.promo_code.code if r.promo_code_id else ''
        amount = ''
        pay = getattr(r, 'payment', None)
        if pay:
            amount = float(pay.net_amount)
        yield [
            full_name, u.username, u.email, tier, r.status,
            'Yes' if r.is_checked_in else 'No', seat, zone, promo, amount,
            r.registered_at.strftime('%Y-%m-%d %H:%M'), str(r.ticket_uuid),
        ]


def attendees_csv_response(event) -> HttpResponse:
    """Return a styled .xlsx workbook of the event's attendees.

    (Name kept for backwards compatibility with the existing view import.)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Attendees'

    header_fill = PatternFill('solid', fgColor='6366F1')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin = Side(style='thin', color='E5E7EB')
    border = Border(bottom=thin)

    # Header row
    ws.append(HEADERS)
    for col, _ in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(vertical='center', horizontal='left')
    ws.row_dimensions[1].height = 22

    # Data rows
    widths = [len(h) for h in HEADERS]
    row_count = 0
    for row in _rows(event):
        ws.append(row)
        row_count += 1
        for i, val in enumerate(row):
            widths[i] = max(widths[i], len(str(val)))
            ws.cell(row=row_count + 1, column=i + 1).border = border

    # Amount paid as currency
    amount_col = HEADERS.index('Amount paid') + 1
    for r in range(2, row_count + 2):
        cell = ws.cell(row=r, column=amount_col)
        if isinstance(cell.value, (int, float)):
            cell.number_format = '$#,##0.00'

    # Auto width (capped) + freeze header + filter
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 10), 42)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(HEADERS))}{row_count + 1}'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_title = ''.join(c if c.isalnum() else '-' for c in event.title)[:40]
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{safe_title}-attendees.xlsx"'
    return response
